"""재무건전성 평가 Excel 내보내기 서비스."""

import io
from datetime import date
from typing import Optional

import openpyxl
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter

from app.database import get_db
from app.services.financial_health import calculate_health


# ─── 색상 상수 ───────────────────────────────────────────────────────────────
GRADE_FILL = {
    "AAA": "D1FAE5", "AA": "DCFCE7", "A": "DBEAFE",
    "BBB": "FEF9C3", "BB": "FFEDD5", "B": "FEE2E2",
}
ITEM_GRADE_FILL = {
    "A (양호)": "D1FAE5", "B (보통)": "DBEAFE",
    "C (주의)": "FEF9C3", "D (위험)": "FEE2E2", "N/A": "F3F4F6",
}
HEADER_FILL  = PatternFill("solid", fgColor="1E3A5F")
SECTION_FILL = PatternFill("solid", fgColor="E2E8F0")
ZEBRA_FILL   = PatternFill("solid", fgColor="F8FAFC")

THIN = Side(style="thin", color="CBD5E1")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

DOMAIN_FILLS = {
    "안전성":   "EFF6FF",
    "수익성":   "F5F3FF",
    "성장성":   "ECFDF5",
    "활동성":   "FFFBEB",
    "현금흐름": "FFF1F2",
}


# ─── 헬퍼 ────────────────────────────────────────────────────────────────────

def _cell(ws, row, col, value="", bold=False, size=11, color="000000",
          fill: Optional[PatternFill] = None, align="left",
          border=True, wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, size=size, color=color,
                  name="맑은 고딕")
    c.alignment = Alignment(horizontal=align, vertical="center",
                             wrap_text=wrap)
    if fill:
        c.fill = fill
    if border:
        c.border = THIN_BORDER
    return c


def _merge(ws, r1, c1, r2, c2, value="", bold=False, size=11,
           color="000000", fill: Optional[PatternFill] = None,
           align="left", wrap=False):
    ws.merge_cells(start_row=r1, start_column=c1,
                   end_row=r2,   end_column=c2)
    c = ws.cell(row=r1, column=c1, value=value)
    c.font = Font(bold=bold, size=size, color=color, name="맑은 고딕")
    c.alignment = Alignment(horizontal=align, vertical="center",
                             wrap_text=wrap)
    if fill:
        c.fill = fill
    c.border = THIN_BORDER
    return c


# ─── Sheet 1: 종합 평가표 ─────────────────────────────────────────────────────

def _build_summary_sheet(ws, company: dict, health: dict):
    # 열 너비
    col_widths = [3, 14, 22, 16, 18, 10, 10, 14, 30]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 8

    # ── 타이틀 ──────────────────────────────────────────────────────────────
    _merge(ws, 2, 2, 2, 9,
           value="거래처 재무건전성 평가표  |  자동차부품업 특화",
           bold=True, size=14, color="FFFFFF",
           fill=HEADER_FILL, align="center")
    ws.row_dimensions[2].height = 32

    # ── 기업 정보 ────────────────────────────────────────────────────────────
    info_rows = [
        ("회사명",       company.get("company_name", "-")),
        ("사업자번호",   company.get("biz_no", "-")),
        ("대표자",       company.get("representatives", "-")),
        ("업종",         company.get("industry", "-")),
        ("주요제품",     company.get("main_product", "-")),
        ("평가기준기간", health.get("period", "-")),
        ("평가일",       str(date.today())),
    ]
    r = 4
    _merge(ws, r, 2, r, 9, value="■ 기업 기본정보",
           bold=True, size=10, fill=SECTION_FILL)
    ws.row_dimensions[r].height = 20
    r += 1
    for label, val in info_rows:
        _cell(ws, r, 2, label, bold=True, size=10,
              fill=PatternFill("solid", fgColor="F1F5F9"))
        _merge(ws, r, 3, r, 9, value=val, size=10)
        ws.row_dimensions[r].height = 18
        r += 1

    r += 1  # 공백

    # ── 종합 등급 ────────────────────────────────────────────────────────────
    grade     = health.get("grade", "-")
    total     = health.get("total_score", 0)
    rec       = health.get("recommendation", "-")
    grade_hex = GRADE_FILL.get(grade, "F3F4F6")

    _merge(ws, r, 2, r, 9, value="■ 종합 평가 결과",
           bold=True, size=10, fill=SECTION_FILL)
    ws.row_dimensions[r].height = 20
    r += 1

    _cell(ws, r, 2, "종합등급", bold=True, size=10,
          fill=PatternFill("solid", fgColor="F1F5F9"))
    _merge(ws, r, 3, r, 4,
           value=grade, bold=True, size=16, align="center",
           fill=PatternFill("solid", fgColor=grade_hex))
    _cell(ws, r, 5, "종합점수", bold=True, size=10,
          fill=PatternFill("solid", fgColor="F1F5F9"))
    _merge(ws, r, 6, r, 7,
           value=f"{total}점 / 100점", bold=True, size=11, align="center",
           fill=PatternFill("solid", fgColor=grade_hex))
    ws.row_dimensions[r].height = 30
    r += 1

    _cell(ws, r, 2, "거래 권고", bold=True, size=10,
          fill=PatternFill("solid", fgColor="F1F5F9"))
    _merge(ws, r, 3, r, 9, value=rec, size=11,
           fill=PatternFill("solid", fgColor=grade_hex))
    ws.row_dimensions[r].height = 22
    r += 2

    # ── 등급 기준표 (소형) ───────────────────────────────────────────────────
    _merge(ws, r, 2, r, 9, value="■ 등급 기준",
           bold=True, size=10, fill=SECTION_FILL)
    ws.row_dimensions[r].height = 20
    r += 1
    grade_refs = [
        ("AAA", "85점↑", "최우량", "✅ 적극 거래 권장"),
        ("AA",  "75~84점", "우량",  "✅ 거래 계속 / 정기 모니터링"),
        ("A",   "65~74점", "양호",  "✅ 거래 계속 / 일부 지표 점검"),
        ("BBB", "55~64점", "보통",  "⚠️ 조건부 거래 / 결제조건 재검토"),
        ("BB",  "45~54점", "주의",  "⚠️ 거래 축소 / 선결제 또는 담보 요구"),
        ("B",   "44점↓",  "위험",  "🚫 거래 재검토 / 신규 수주 중단 검토"),
    ]
    for g, score_range, eval_, action in grade_refs:
        fill = PatternFill("solid", fgColor=GRADE_FILL.get(g, "F3F4F6"))
        _cell(ws, r, 2, g,           bold=True, size=10, fill=fill, align="center")
        _cell(ws, r, 3, score_range, size=10, fill=fill, align="center")
        _cell(ws, r, 4, eval_,       bold=True, size=10, fill=fill, align="center")
        _merge(ws, r, 5, r, 9, value=action, size=10, fill=fill)
        ws.row_dimensions[r].height = 18
        r += 1

    r += 1

    # ── 항목별 세부 점수표 ────────────────────────────────────────────────────
    _merge(ws, r, 2, r, 9, value="■ 항목별 세부 평가",
           bold=True, size=10, fill=SECTION_FILL)
    ws.row_dimensions[r].height = 20
    r += 1

    # 헤더
    headers = ["영역", "지표", "당기값", "단위", "벤치마크", "배점", "점수", "달성률", "항목등급"]
    for ci, h in enumerate(headers, 2):
        _cell(ws, r, ci, h, bold=True, size=10, color="FFFFFF",
              fill=HEADER_FILL, align="center")
    ws.row_dimensions[r].height = 20
    r += 1

    # 데이터 행
    for di, domain in enumerate(health.get("domains", [])):
        domain_fill = PatternFill("solid", fgColor=DOMAIN_FILLS.get(domain["name"], "FFFFFF"))
        items = domain["items"]
        start_r = r
        for ii, item in enumerate(items):
            if ii == 0:
                _merge(ws, r, 2, r + len(items) - 1, 2,
                       value=domain["name"], bold=True, size=10,
                       fill=domain_fill, align="center")

            val_display = f"{item['value']:,}" if item['value'] is not None else "-"
            item_fill = PatternFill("solid", fgColor=ITEM_GRADE_FILL.get(item["item_grade"], "FFFFFF"))
            pct = f"{item['score']/item['max_score']*100:.0f}%" if item["max_score"] > 0 else "-"

            zebra = ZEBRA_FILL if (di + ii) % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
            _cell(ws, r, 3, item["label"],      size=10, fill=zebra)
            _cell(ws, r, 4, val_display,        size=10, fill=zebra, align="right")
            _cell(ws, r, 5, item["unit"],        size=10, fill=zebra, align="center")
            _cell(ws, r, 6, item["benchmark"],   size=10, fill=zebra, align="center")
            _cell(ws, r, 7, item["max_score"],   size=10, fill=zebra, align="center")
            _cell(ws, r, 8, item["score"],       bold=True, size=10, fill=item_fill, align="center")
            _cell(ws, r, 9, pct,                 size=10, fill=item_fill, align="center")
            _cell(ws, r, 10, item["item_grade"], bold=True, size=10, fill=item_fill, align="center")
            ws.row_dimensions[r].height = 18
            r += 1

        # 도메인 소계
        _cell(ws, r, 2, "",                          fill=domain_fill)
        _cell(ws, r, 3, f"{domain['name']} 소계",    bold=True, size=10, fill=domain_fill)
        _cell(ws, r, 4, "",  fill=domain_fill)
        _cell(ws, r, 5, "",  fill=domain_fill)
        _cell(ws, r, 6, "",  fill=domain_fill)
        _cell(ws, r, 7, domain["max_score"],  bold=True, size=10, fill=domain_fill, align="center")
        _cell(ws, r, 8, domain["score"],      bold=True, size=10, fill=domain_fill, align="center")
        d_pct = f"{domain['score']/domain['max_score']*100:.0f}%" if domain["max_score"] > 0 else "-"
        _cell(ws, r, 9, d_pct,  bold=True, size=10, fill=domain_fill, align="center")
        _cell(ws, r, 10, "",    fill=domain_fill)
        ws.row_dimensions[r].height = 18
        r += 1

    # 합계
    total_fill = PatternFill("solid", fgColor=GRADE_FILL.get(grade, "F3F4F6"))
    _merge(ws, r, 2, r, 7, value="합  계", bold=True, size=11,
           fill=total_fill, align="center")
    _cell(ws, r, 8, total,  bold=True, size=12, fill=total_fill, align="center")
    _cell(ws, r, 9, f"{total}%", bold=True, size=11, fill=total_fill, align="center")
    _merge(ws, r, 10, r, 10, value=f"{grade} ({health.get('recommendation','')})",
           bold=True, size=10, fill=total_fill)
    ws.row_dimensions[r].height = 24

    r += 2

    # ── 자동차부품업 모니터링 항목 ────────────────────────────────────────────
    _merge(ws, r, 2, r, 9, value="■ 자동차부품업 특화 모니터링 항목",
           bold=True, size=10, fill=SECTION_FILL)
    ws.row_dimensions[r].height = 20
    r += 1
    monitors = [
        "① OEM 납품 집중도 — 특정 완성차사 매출 의존도 50% 초과 시 리스크 상승",
        "② 철강재 원가 연동 여부 — 원자재 가격 변동 계약 조건 확인 (변동 vs 고정)",
        "③ 단기차입금 비중 — 차입금 중 단기(1년 이내) 비율 60% 초과 시 차환 리스크",
        "④ 영업현금흐름 vs 순이익 괴리 — 현금흐름이 이익보다 낮으면 운전자본 압박 가능성",
        "⑤ 전기차(EV) 부품 대응 여부 — 내연기관 특화 부품사의 중장기 사업전환 계획 확인",
        "⑥ 결제방법 변화 — 현금결제→어음/외상 전환은 자금난 신호일 수 있음",
    ]
    for m in monitors:
        _merge(ws, r, 2, r, 9, value=m, size=10,
               fill=PatternFill("solid", fgColor="FFFBEB"), wrap=True)
        ws.row_dimensions[r].height = 20
        r += 1

    ws.freeze_panes = "B4"
    ws.sheet_view.showGridLines = False


# ─── Sheet 2: 재무 원시데이터 ─────────────────────────────────────────────────

def _build_rawdata_sheet(ws, company_id: int, company: dict):
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 32
    col_widths_raw = [3, 32, 16, 16, 16]
    for i, w in enumerate(col_widths_raw, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    with get_db() as conn:
        periods = [r["period"] for r in conn.execute(
            "SELECT DISTINCT period FROM report_values "
            "WHERE import_id=? AND period != '-' ORDER BY period",
            (company_id,),
        ).fetchall()]

        sections_to_show = ["재무상태표", "손익계산서", "포괄손익계산서", "현금흐름분석"]
        all_rows = []
        for sec in sections_to_show:
            rows = conn.execute(
                "SELECT section, metric, period, value_raw, value_num, unit "
                "FROM report_values "
                "WHERE import_id=? AND section=? AND submetric IS NULL AND category IS NULL "
                "ORDER BY row_no",
                (company_id, sec),
            ).fetchall()
            if rows:
                all_rows.append((sec, rows))

        unit_map = {}
        for sec, _ in all_rows:
            u = conn.execute(
                "SELECT unit FROM report_values WHERE import_id=? AND section=? AND unit IS NOT NULL LIMIT 1",
                (company_id, sec),
            ).fetchone()
            unit_map[sec] = u["unit"] if u else ""

    # 열 너비 (period 수에 따라)
    for i, p in enumerate(periods, 3):
        ws.column_dimensions[get_column_letter(i)].width = 16

    r = 1
    ws.row_dimensions[r].height = 8
    r += 1

    _merge(ws, r, 1, r, 2 + len(periods),
           value=f"재무 원시데이터  |  {company.get('company_name', '')}",
           bold=True, size=13, color="FFFFFF",
           fill=HEADER_FILL, align="center")
    ws.row_dimensions[r].height = 30
    r += 2

    for sec, rows in all_rows:
        unit_label = f"  (단위: {unit_map[sec]})" if unit_map[sec] else ""
        _merge(ws, r, 1, r, 2 + len(periods),
               value=f"■ {sec}{unit_label}",
               bold=True, size=10, fill=SECTION_FILL)
        ws.row_dimensions[r].height = 20
        r += 1

        # 헤더 행
        _cell(ws, r, 1, "#",      bold=True, size=9, fill=HEADER_FILL,
              color="FFFFFF", align="center")
        _cell(ws, r, 2, "항목명", bold=True, size=9, fill=HEADER_FILL,
              color="FFFFFF")
        for ci, p in enumerate(periods, 3):
            _cell(ws, r, ci, p, bold=True, size=9, fill=HEADER_FILL,
                  color="FFFFFF", align="center")
        ws.row_dimensions[r].height = 18
        r += 1

        # pivot: metric -> {period: value_raw}
        seen: list = []
        pivot: dict = {}
        for row in rows:
            m = row["metric"]
            if m not in pivot:
                pivot[m] = {}
                seen.append(m)
            pivot[m][row["period"]] = row["value_raw"]

        for idx, m in enumerate(seen):
            fill = ZEBRA_FILL if idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
            is_total = any(kw in m for kw in ["총계", "합계", "소계", "순이익", "영업이익"])
            row_fill = PatternFill("solid", fgColor="EFF6FF") if is_total else fill
            _cell(ws, r, 1, idx + 1, size=9, fill=row_fill, align="center")
            _cell(ws, r, 2, m, bold=is_total, size=9, fill=row_fill)
            for ci, p in enumerate(periods, 3):
                val = pivot[m].get(p, "")
                _cell(ws, r, ci, val, size=9, fill=row_fill, align="right")
            ws.row_dimensions[r].height = 16
            r += 1

        r += 1

    ws.freeze_panes = "C4"
    ws.sheet_view.showGridLines = False


# ─── 메인 함수 ───────────────────────────────────────────────────────────────

def export_health_excel(company_id: int) -> io.BytesIO:
    """기업 재무건전성 평가 Excel 파일을 BytesIO로 반환."""
    health = calculate_health(company_id)

    with get_db() as conn:
        row = conn.execute(
            "SELECT company_name, representatives, biz_no, report_date, industry, main_product "
            "FROM report_imports WHERE id=?",
            (company_id,),
        ).fetchone()
    company = dict(row) if row else {}

    wb = openpyxl.Workbook()

    # Sheet 1
    ws1 = wb.active
    ws1.title = "종합 평가표"
    _build_summary_sheet(ws1, company, health)

    # Sheet 2
    ws2 = wb.create_sheet("재무 원시데이터")
    _build_rawdata_sheet(ws2, company_id, company)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, company.get("company_name", str(company_id))
