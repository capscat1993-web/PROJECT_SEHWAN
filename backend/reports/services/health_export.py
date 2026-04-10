"""재무건전성 평가 Excel 내보내기 서비스."""

import io
from datetime import date
from typing import Optional

import openpyxl
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter

from reports.db import get_db
from reports.services.financial_health import calculate_health, get_operating_cashflow


# ─── 색상 상수 ───────────────────────────────────────────────────────────────
GRADE_FILL = {
    "AAA": "D1FAE5", "AA": "DCFCE7", "A": "DBEAFE",
    "BBB": "FEF9C3", "BB": "FFEDD5", "B": "FEE2E2",
}
ITEM_GRADE_FILL = {
    "A (양호)": "D1FAE5", "B (보통)": "DBEAFE",
    "C (주의)": "FEF9C3", "D (위험)": "FEE2E2", "N/A": "F3F4F6",
}
HEADER_FILL  = PatternFill("solid", fgColor="1E3A5F")
SECTION_FILL = PatternFill("solid", fgColor="E2E8F0")
ZEBRA_FILL   = PatternFill("solid", fgColor="F8FAFC")

THIN = Side(style="thin", color="CBD5E1")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

DOMAIN_FILLS = {
    "안전성":   "EFF6FF",
    "수익성":   "F5F3FF",
    "성장성":   "ECFDF5",
    "활동성":   "FFFBEB",
    "현금흐름": "FFF1F2",
}


# ─── 헬퍼 ────────────────────────────────────────────────────────────────────

def _cell(ws, row, col, value="", bold=False, size=11, color="000000",
          fill: Optional[PatternFill] = None, align="left",
          border=True, wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, size=size, color=color,
                  name="맑은 고딕")
    c.alignment = Alignment(horizontal=align, vertical="center",
                             wrap_text=wrap)
    if fill:
        c.fill = fill
    if border:
        c.border = THIN_BORDER
    return c


def _merge(ws, r1, c1, r2, c2, value="", bold=False, size=11,
           color="000000", fill: Optional[PatternFill] = None,
           align="left", wrap=False):
    ws.merge_cells(start_row=r1, start_column=c1,
                   end_row=r2,   end_column=c2)
    c = ws.cell(row=r1, column=c1, value=value)
    c.font = Font(bold=bold, size=size, color=color, name="맑은 고딕")
    c.alignment = Alignment(horizontal=align, vertical="center",
                             wrap_text=wrap)
    if fill:
        c.fill = fill
    c.border = THIN_BORDER
    return c


# ─── Sheet 1: 종합 평가표 ─────────────────────────────────────────────────────

def _build_summary_sheet(ws, company: dict, health: dict):
    # 열 너비
    col_widths = [3, 14, 22, 16, 18, 10, 10, 14, 30]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 8

    # ── 타이틀 ──────────────────────────────────────────────────────────────
    _merge(ws, 2, 2, 2, 9,
           value="거래처 재무건전성 평가표  |  자동차부품업 특화",
           bold=True, size=14, color="FFFFFF",
           fill=HEADER_FILL, align="center")
    ws.row_dimensions[2].height = 32

    # ── 기업 정보 ────────────────────────────────────────────────────────────
    info_rows = [
        ("회사명",       company.get("company_name", "-")),
        ("사업자번호",   company.get("biz_no", "-")),
        ("대표자",       company.get("representatives", "-")),
        ("업종",         company.get("industry", "-")),
        ("주요제품",     company.get("main_product", "-")),
        ("평가기준기간", health.get("period", "-")),
        ("평가일",       str(date.today())),
    ]
    r = 4
    _merge(ws, r, 2, r, 9, value="■ 기업 기본정보",
           bold=True, size=10, fill=SECTION_FILL)
    ws.row_dimensions[r].height = 20
    r += 1
    for label, val in info_rows:
        _cell(ws, r, 2, label, bold=True, size=10,
              fill=PatternFill("solid", fgColor="F1F5F9"))
        _merge(ws, r, 3, r, 9, value=val, size=10)
        ws.row_dimensions[r].height = 18
        r += 1

    r += 1  # 공백

    # ── 종합 등급 ────────────────────────────────────────────────────────────
    grade     = health.get("grade", "-")
    total     = health.get("total_score", 0)
    rec       = health.get("recommendation", "-")
    grade_note = health.get("grade_note", "")
    data_note = health.get("data_note", "")
    rec_parts = [rec]
    if grade_note:
        rec_parts.append(grade_note)
    if data_note:
        rec_parts.append(data_note)
    rec_display = "\n".join(part for part in rec_parts if part)
    grade_hex = GRADE_FILL.get(grade, "F3F4F6")

    _merge(ws, r, 2, r, 9, value="■ 종합 평가 결과",
           bold=True, size=10, fill=SECTION_FILL)
    ws.row_dimensions[r].height = 20
    r += 1

    _cell(ws, r, 2, "종합등급", bold=True, size=10,
          fill=PatternFill("solid", fgColor="F1F5F9"))
    _merge(ws, r, 3, r, 4,
           value=grade, bold=True, size=16, align="center",
           fill=PatternFill("solid", fgColor=grade_hex))
    _cell(ws, r, 5, "종합점수", bold=True, size=10,
          fill=PatternFill("solid", fgColor="F1F5F9"))
    _merge(ws, r, 6, r, 7,
           value=f"{total}점 / 100점", bold=True, size=11, align="center",
           fill=PatternFill("solid", fgColor=grade_hex))
    ws.row_dimensions[r].height = 30
    r += 1

    _cell(ws, r, 2, "거래 권고", bold=True, size=10,
          fill=PatternFill("solid", fgColor="F1F5F9"))
    _merge(ws, r, 3, r, 9, value=rec_display, size=11, wrap=True,
           fill=PatternFill("solid", fgColor=grade_hex))
    ws.row_dimensions[r].height = 44 if grade_note or data_note else 22
    r += 2

    # ── 등급 기준표 (소형) ───────────────────────────────────────────────────
    _merge(ws, r, 2, r, 9, value="■ 등급 기준",
           bold=True, size=10, fill=SECTION_FILL)
    ws.row_dimensions[r].height = 20
    r += 1
    grade_refs = [
        ("AAA", "85점↑", "최우량", "✅ 적극 거래 권장"),
        ("AA",  "75~84점", "우량",  "✅ 거래 계속 / 정기 모니터링"),
        ("A",   "65~74점", "양호",  "✅ 거래 계속 / 일부 지표 점검"),
        ("BBB", "55~64점", "보통",  "⚠️ 조건부 거래 / 결제조건 재검토"),
        ("BB",  "45~54점", "주의",  "⚠️ 거래 축소 / 선결제 또는 담보 요구"),
        ("B",   "44점↓",  "위험",  "🚫 거래 재검토 / 신규 수주 중단 검토"),
    ]
    for g, score_range, eval_, action in grade_refs:
        fill = PatternFill("solid", fgColor=GRADE_FILL.get(g, "F3F4F6"))
        _cell(ws, r, 2, g,           bold=True, size=10, fill=fill, align="center")
        _cell(ws, r, 3, score_range, size=10, fill=fill, align="center")
        _cell(ws, r, 4, eval_,       bold=True, size=10, fill=fill, align="center")
        _merge(ws, r, 5, r, 9, value=action, size=10, fill=fill)
        ws.row_dimensions[r].height = 18
        r += 1

    r += 1

    # ── 항목별 세부 점수표 ────────────────────────────────────────────────────
    _merge(ws, r, 2, r, 9, value="■ 항목별 세부 평가",
           bold=True, size=10, fill=SECTION_FILL)
    ws.row_dimensions[r].height = 20
    r += 1

    # 헤더
    headers = ["영역", "지표", "당기값", "단위", "벤치마크", "배점", "점수", "달성률", "항목등급"]
    for ci, h in enumerate(headers, 2):
        _cell(ws, r, ci, h, bold=True, size=10, color="FFFFFF",
              fill=HEADER_FILL, align="center")
    ws.row_dimensions[r].height = 20
    r += 1

    # 데이터 행
    for di, domain in enumerate(health.get("domains", [])):
        domain_fill = PatternFill("solid", fgColor=DOMAIN_FILLS.get(domain["name"], "FFFFFF"))
        items = domain["items"]
        start_r = r
        for ii, item in enumerate(items):
            if ii == 0:
                _merge(ws, r, 2, r + len(items) - 1, 2,
                       value=domain["name"], bold=True, size=10,
                       fill=domain_fill, align="center")

            val_display = f"{item['value']:,}" if item['value'] is not None else "-"
            item_fill = PatternFill("solid", fgColor=ITEM_GRADE_FILL.get(item["item_grade"], "FFFFFF"))
            pct = f"{item['score']/item['max_score']*100:.0f}%" if item["max_score"] > 0 else "-"

            zebra = ZEBRA_FILL if (di + ii) % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
            _cell(ws, r, 3, item["label"],      size=10, fill=zebra)
            _cell(ws, r, 4, val_display,        size=10, fill=zebra, align="right")
            _cell(ws, r, 5, item["unit"],        size=10, fill=zebra, align="center")
            _cell(ws, r, 6, item["benchmark"],   size=10, fill=zebra, align="center")
            _cell(ws, r, 7, item["max_score"],   size=10, fill=zebra, align="center")
            _cell(ws, r, 8, item["score"],       bold=True, size=10, fill=item_fill, align="center")
            _cell(ws, r, 9, pct,                 size=10, fill=item_fill, align="center")
            _cell(ws, r, 10, item["item_grade"], bold=True, size=10, fill=item_fill, align="center")
            ws.row_dimensions[r].height = 18
            r += 1

        # 도메인 소계
        _cell(ws, r, 2, "",                          fill=domain_fill)
        _cell(ws, r, 3, f"{domain['name']} 소계",    bold=True, size=10, fill=domain_fill)
        _cell(ws, r, 4, "",  fill=domain_fill)
        _cell(ws, r, 5, "",  fill=domain_fill)
        _cell(ws, r, 6, "",  fill=domain_fill)
        _cell(ws, r, 7, domain["max_score"],  bold=True, size=10, fill=domain_fill, align="center")
        _cell(ws, r, 8, domain["score"],      bold=True, size=10, fill=domain_fill, align="center")
        d_pct = f"{domain['score']/domain['max_score']*100:.0f}%" if domain["max_score"] > 0 else "-"
        _cell(ws, r, 9, d_pct,  bold=True, size=10, fill=domain_fill, align="center")
        _cell(ws, r, 10, "",    fill=domain_fill)
        ws.row_dimensions[r].height = 18
        r += 1

    # 합계
    total_fill = PatternFill("solid", fgColor=GRADE_FILL.get(grade, "F3F4F6"))
    _merge(ws, r, 2, r, 7, value="합  계", bold=True, size=11,
           fill=total_fill, align="center")
    _cell(ws, r, 8, total,  bold=True, size=12, fill=total_fill, align="center")
    _cell(ws, r, 9, f"{total}%", bold=True, size=11, fill=total_fill, align="center")
    _merge(ws, r, 10, r, 10, value=f"{grade} ({health.get('recommendation','')})",
           bold=True, size=10, fill=total_fill)
    ws.row_dimensions[r].height = 24

    r += 2

    # ── 자동차부품업 모니터링 항목 ────────────────────────────────────────────
    _merge(ws, r, 2, r, 9, value="■ 자동차부품업 특화 모니터링 항목",
           bold=True, size=10, fill=SECTION_FILL)
    ws.row_dimensions[r].height = 20
    r += 1
    monitors = [
        "① OEM 납품 집중도 — 특정 완성차사 매출 의존도 50% 초과 시 리스크 상승",
        "② 철강재 원가 연동 여부 — 원자재 가격 변동 계약 조건 확인 (변동 vs 고정)",
        "③ 단기차입금 비중 — 차입금 중 단기(1년 이내) 비율 60% 초과 시 차환 리스크",
        "④ 영업현금흐름 vs 순이익 괴리 — 현금흐름이 이익보다 낮으면 운전자본 압박 가능성",
        "⑤ 전기차(EV) 부품 대응 여부 — 내연기관 특화 부품사의 중장기 사업전환 계획 확인",
        "⑥ 결제방법 변화 — 현금결제→어음/외상 전환은 자금난 신호일 수 있음",
    ]
    for m in monitors:
        _merge(ws, r, 2, r, 9, value=m, size=10,
               fill=PatternFill("solid", fgColor="FFFBEB"), wrap=True)
        ws.row_dimensions[r].height = 20
        r += 1

    ws.freeze_panes = "B4"
    ws.sheet_view.showGridLines = False


# ─── Sheet 2: 재무비율 분석 ──────────────────────────────────────────────────

# (영역, 지표명, DB섹션, DB메트릭, 단위, 방향, 벤치마크, 등급기준, 설명)
# 방향: "↑" = 높을수록 좋음, "↓" = 낮을수록 좋음
# 활동성 회전율은 DB에서 가져온 뒤 365/값으로 일수 변환
_RATIO_INDICATORS = [
    ("안전성", "유동비율",         "안정성지표",   "유동비율(%)",          "%",      "↑", "100% 이상", "≥150% 양호 / 100~150% 보통 / <100% 위험", "단기채무 대비 유동자산 비율"),
    ("안전성", "부채비율",         "안정성지표",   "부채비율(%)",          "%",      "↓", "150% 이하", "≤100% 양호 / 100~200% 보통 / >200% 위험", "타인자본/자기자본"),
    ("안전성", "자기자본비율",     "안정성지표",   "자기자본비율(%)",       "%",      "↑", "30% 이상",  "≥40% 양호 / 25~40% 보통 / <25% 위험", "자기자본/총자산"),
    ("안전성", "차입금의존도",     "안정성지표",   "차입금의존도(%)",       "%",      "↓", "30% 이하",  "≤20% 양호 / 20~40% 보통 / >40% 위험", "차입금/총자산"),
    ("안전성", "이자보상배율",     "수익성지표",   "이자보상배율(배)",       "배",     "↑", "3배 이상",  "≥5배 양호 / 2~5배 보통 / <2배 위험", "영업이익/이자비용"),
    ("수익성", "영업이익률",       "수익성지표",   "매출액영업이익률(%)",   "%",      "↑", "5% 이상",   "≥7% 양호 / 3~7% 보통 / <3% 위험", "영업이익/매출액"),
    ("수익성", "순이익률",         "수익성지표",   "매출액순이익률(%)",     "%",      "↑", "3% 이상",   "≥5% 양호 / 1~5% 보통 / <1% 위험", "당기순이익/매출액"),
    ("수익성", "ROA",              "수익성지표",   "총자본순이익률(%)",     "%",      "↑", "3% 이상",   "≥5% 양호 / 2~5% 보통 / <2% 위험", "순이익/총자산"),
    ("수익성", "ROE",              "수익성지표",   "자기자본순이익률(%)",   "%",      "↑", "8% 이상",   "≥10% 양호 / 5~10% 보통 / <5% 위험", "순이익/자기자본"),
    ("성장성", "매출액증가율",     "주요재무지표", "매출액증가율(%)",       "%",      "↑", "5% 이상",   "≥10% 양호 / 0~10% 보통 / <0% 위험", "전년비 매출 성장률"),
    ("성장성", "총자산증가율",     "주요재무지표", "총자산증가율(%)",       "%",      "↑", "3% 이상",   "≥5% 양호 / 0~5% 보통 / <0% 위험", "전년비 총자산 성장률"),
    ("활동성", "매출채권회전일수", "활동성지표",   "매출채권회전율(회)",    "일",     "↓", "60일 이하", "≤45일 양호 / 45~90일 보통 / >90일 위험", "365/매출채권회전율"),
    ("활동성", "재고자산회전일수", "활동성지표",   "재고자산회전율(회)",    "일",     "↓", "45일 이하", "≤30일 양호 / 30~60일 보통 / >60일 위험", "365/재고자산회전율"),
    ("활동성", "매입채무회전일수", "활동성지표",   "매입채무회전율(회)",    "일",     "↑", "45~60일",   "결제조건 준수 여부 확인", "365/매입채무회전율"),
    ("현금흐름", "영업현금흐름",   "현금흐름분석", "영업활동 현금흐름",     "백만원", "↑", "양(+)값",   "연속 양(+) 양호 / 음(-) 주의", "영업활동에서 창출한 현금"),
]

_TURNOVER_TO_DAYS = {"매출채권회전일수", "재고자산회전일수", "매입채무회전일수"}

_RATIO_TEMPLATE_EXTRA_DESCRIPTIONS = {
    "유동비율": "단기 채무 대응능력. 1이상이면 단기부채를 유동자산으로 상환 가능",
    "부채비율": "자본 대비 부채 비율. 낮을수록 재무구조 안정적",
    "자기자본비율": "총자산 중 자기자본 비중",
    "차입금의존도": "총자산 대비 금융부채 비율",
    "이자보상배율": "영업이익으로 이자를 몇 배 갚을 수 있는지",
    "영업이익률": "본업 수익성 핵심 지표",
    "순이익률": "최종 수익성",
    "ROA": "총자산 대비 순이익률",
    "ROE": "자기자본 대비 순이익률",
    "매출액증가율": "매출 성장 모멘텀",
    "총자산증가율": "사업 규모 확대 여부",
    "매출채권회전일수": "매출채권 회수 속도. 길수록 자금회수 지연",
    "재고자산회전일수": "재고 소화 속도. 길수록 운전자본 압박",
    "매입채무회전일수": "철강재 결제조건과 비교 필요",
    "영업현금흐름": "영업활동으로 창출한 실질 현금. 순이익보다 중요",
}


def _build_ratio_sheet(ws, company_id: int, company: dict):
    """Sheet 3: 재무비율 분석 — 지표별 3개년 추이 + 벤치마크/등급기준/설명."""

    # ── 열 너비 ─────────────────────────────────────────────────────────────
    col_specs = [
        ("A", 4), ("B", 10), ("C", 18),
        ("D", 14), ("E", 14), ("F", 14),
        ("G", 14), ("H", 10), ("I", 38), ("J", 30), ("K", 48),
    ]
    for col_letter, width in col_specs:
        ws.column_dimensions[col_letter].width = width

    # ── DB 조회 ──────────────────────────────────────────────────────────────
    with get_db() as conn:
        periods = [r["period"] for r in conn.execute(
            "SELECT DISTINCT period FROM report_values "
            "WHERE import_id=? AND period != '-' ORDER BY period",
            (company_id,),
        ).fetchall()]

        def get_val(section, metric, period):
            if metric == "영업활동 현금흐름":
                return get_operating_cashflow(conn, company_id, period)
            row = conn.execute(
                "SELECT value_num FROM report_values "
                "WHERE import_id=? AND section=? AND metric=? AND period=? "
                "AND value_num IS NOT NULL ORDER BY id LIMIT 1",
                (company_id, section, metric, period),
            ).fetchone()
            return row["value_num"] if row else None

        # 지표별 기간값 수집
        indicator_data = []
        for ind in _RATIO_INDICATORS:
            domain, label, sec, metric, unit, direction, benchmark, grade_criteria, desc = ind
            extra_desc = _RATIO_TEMPLATE_EXTRA_DESCRIPTIONS.get(label, desc)
            vals = []
            for p in periods:
                v = get_val(sec, metric, p)
                if label in _TURNOVER_TO_DAYS and v and v > 0:
                    v = round(365 / v, 1)
                elif v is not None:
                    v = round(v, 2)
                vals.append(v)
            indicator_data.append((domain, label, unit, direction, benchmark, grade_criteria, desc, extra_desc, vals))

    # 최근 3개 기간만 사용
    if len(periods) >= 3:
        display_periods = periods[-3:]
    else:
        display_periods = periods
    # 데이터도 최근 3개 슬라이싱
    period_count = len(display_periods)
    data_offset = len(periods) - period_count  # 앞 기간을 잘라낼 offset

    def _fmt(v):
        if v is None:
            return "-"
        return f"{v:,.2f}".rstrip("0").rstrip(".")

    def _wrapped_height(text: object, chars_per_line: int, line_height: int = 16) -> int:
        value = "" if text is None else str(text)
        if not value:
            return line_height
        lines = 0
        for chunk in value.splitlines() or [""]:
            chunk_len = max(1, len(chunk))
            lines += max(1, (chunk_len + chars_per_line - 1) // chars_per_line)
        return lines * line_height

    def _trend(vals, direction):
        """FY-1 → FY 변화 기반 추세."""
        trimmed = [v for v in vals[-2:] if v is not None]
        if len(trimmed) < 2:
            return "–"
        delta = trimmed[-1] - trimmed[-2]
        if abs(delta) < 0.01:
            return "→ 유지"
        improved = (delta > 0 and direction == "↑") or (delta < 0 and direction == "↓")
        return "▲ 개선" if improved else "▼ 악화"

    TREND_FILL = {
        "▲ 개선": PatternFill("solid", fgColor="D1FAE5"),
        "▼ 악화": PatternFill("solid", fgColor="FEE2E2"),
        "→ 유지": PatternFill("solid", fgColor="F3F4F6"),
        "–":      PatternFill("solid", fgColor="F3F4F6"),
    }

    # ── 타이틀 ───────────────────────────────────────────────────────────────
    r = 1
    ws.row_dimensions[r].height = 8
    r += 1
    _merge(ws, r, 1, r, 11,
           value=f"재무비율 분석  |  {company.get('company_name', '')}",
           bold=True, size=13, color="FFFFFF",
           fill=HEADER_FILL, align="center")
    ws.row_dimensions[r].height = 30
    r += 1

    # 기간 레이블 (FY-2 / FY-1 / FY)
    period_labels = []
    for i, p in enumerate(display_periods):
        offset = period_count - 1 - i
        period_labels.append(f"FY-{offset}" if offset > 0 else f"FY ({p})")

    # ── 헤더 ─────────────────────────────────────────────────────────────────
    r += 1
    headers = ["No.", "영역", "지표"] + period_labels + ["벤치마크", "추세", "등급기준", "설명", "부가설명"]
    for ci, h in enumerate(headers, 1):
        _cell(ws, r, ci, h, bold=True, size=10, color="FFFFFF",
              fill=HEADER_FILL, align="center")
    ws.row_dimensions[r].height = 30
    r += 1

    # ── 데이터 행 ────────────────────────────────────────────────────────────
    current_domain = None
    domain_start_r = r
    domain_rows: list = []  # (domain, start, end)용
    seq = 1

    for idx, (domain, label, unit, direction, benchmark, grade_criteria, desc, extra_desc, all_vals) in enumerate(indicator_data):
        vals = all_vals[data_offset:] if data_offset > 0 else all_vals
        # 길이 맞추기
        while len(vals) < period_count:
            vals = [None] + vals

        trend = _trend(vals, direction)
        domain_fill = PatternFill("solid", fgColor=DOMAIN_FILLS.get(domain, "FFFFFF"))
        trend_fill = TREND_FILL.get(trend, PatternFill("solid", fgColor="F3F4F6"))
        zebra = ZEBRA_FILL if idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")

        _cell(ws, r, 1, seq,     size=9, fill=zebra, align="center")
        _cell(ws, r, 2, domain,  size=9, fill=domain_fill, align="center", bold=True)
        _cell(ws, r, 3, f"{label}  ({unit})", size=9, fill=zebra)

        for ci, v in enumerate(vals, 4):
            _cell(ws, r, ci, _fmt(v), size=9, fill=zebra, align="right")

        _cell(ws, r, 4 + period_count,     benchmark,      size=9, fill=zebra, align="center")
        _cell(ws, r, 5 + period_count,     trend,          size=9, fill=trend_fill, align="center", bold=True)
        _cell(ws, r, 6 + period_count,     grade_criteria, size=9, fill=zebra, wrap=True)
        _cell(ws, r, 7 + period_count,     desc,           size=9, fill=zebra, wrap=True)
        _cell(ws, r, 8 + period_count,     extra_desc,     size=9, fill=zebra, wrap=True)
        ws.row_dimensions[r].height = max(
            22,
            _wrapped_height(grade_criteria, 24),
            _wrapped_height(desc, 20),
            _wrapped_height(extra_desc, 34),
        )
        r += 1
        seq += 1

    ws.freeze_panes = "D5"
    ws.sheet_view.showGridLines = False


# ─── Sheet 3: 재무 원시데이터 ─────────────────────────────────────────────────

def _build_rawdata_sheet(ws, company_id: int, company: dict):
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 32
    col_widths_raw = [3, 32, 16, 16, 16]
    for i, w in enumerate(col_widths_raw, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    with get_db() as conn:
        periods = [r["period"] for r in conn.execute(
            "SELECT DISTINCT period FROM report_values "
            "WHERE import_id=? AND period != '-' ORDER BY period",
            (company_id,),
        ).fetchall()]

        sections_to_show = ["재무상태표", "손익계산서", "포괄손익계산서", "현금흐름분석"]
        all_rows = []
        for sec in sections_to_show:
            rows = conn.execute(
                "SELECT section, metric, period, value_raw, value_num, unit "
                "FROM report_values "
                "WHERE import_id=? AND section=? AND submetric IS NULL AND category IS NULL "
                "ORDER BY row_no",
                (company_id, sec),
            ).fetchall()
            if rows:
                all_rows.append((sec, rows))

        unit_map = {}
        for sec, _ in all_rows:
            u = conn.execute(
                "SELECT unit FROM report_values WHERE import_id=? AND section=? AND unit IS NOT NULL LIMIT 1",
                (company_id, sec),
            ).fetchone()
            unit_map[sec] = u["unit"] if u else ""

    # 열 너비 (period 수에 따라)
    for i, p in enumerate(periods, 3):
        ws.column_dimensions[get_column_letter(i)].width = 16

    r = 1
    ws.row_dimensions[r].height = 8
    r += 1

    _merge(ws, r, 1, r, 2 + len(periods),
           value=f"재무 원시데이터  |  {company.get('company_name', '')}",
           bold=True, size=13, color="FFFFFF",
           fill=HEADER_FILL, align="center")
    ws.row_dimensions[r].height = 30
    r += 2

    for sec, rows in all_rows:
        unit_label = f"  (단위: {unit_map[sec]})" if unit_map[sec] else ""
        _merge(ws, r, 1, r, 2 + len(periods),
               value=f"■ {sec}{unit_label}",
               bold=True, size=10, fill=SECTION_FILL)
        ws.row_dimensions[r].height = 20
        r += 1

        # 헤더 행
        _cell(ws, r, 1, "#",      bold=True, size=9, fill=HEADER_FILL,
              color="FFFFFF", align="center")
        _cell(ws, r, 2, "항목명", bold=True, size=9, fill=HEADER_FILL,
              color="FFFFFF")
        for ci, p in enumerate(periods, 3):
            _cell(ws, r, ci, p, bold=True, size=9, fill=HEADER_FILL,
                  color="FFFFFF", align="center")
        ws.row_dimensions[r].height = 18
        r += 1

        # pivot: metric -> {period: value_raw}
        seen: list = []
        pivot: dict = {}
        for row in rows:
            m = row["metric"]
            if m not in pivot:
                pivot[m] = {}
                seen.append(m)
            pivot[m][row["period"]] = row["value_raw"]

        for idx, m in enumerate(seen):
            fill = ZEBRA_FILL if idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
            is_total = any(kw in m for kw in ["총계", "합계", "소계", "순이익", "영업이익"])
            row_fill = PatternFill("solid", fgColor="EFF6FF") if is_total else fill
            _cell(ws, r, 1, idx + 1, size=9, fill=row_fill, align="center")
            _cell(ws, r, 2, m, bold=is_total, size=9, fill=row_fill)
            for ci, p in enumerate(periods, 3):
                val = pivot[m].get(p, "")
                _cell(ws, r, ci, val, size=9, fill=row_fill, align="right")
            ws.row_dimensions[r].height = 16
            r += 1

        r += 1

    ws.freeze_panes = "C4"
    ws.sheet_view.showGridLines = False


# ─── 메인 함수 ───────────────────────────────────────────────────────────────

def export_health_excel(company_id: int) -> io.BytesIO:
    """기업 재무건전성 평가 Excel 파일을 BytesIO로 반환."""
    health = calculate_health(company_id)

    with get_db() as conn:
        row = conn.execute(
            "SELECT company_name, representatives, biz_no, report_date, industry, main_product "
            "FROM report_imports WHERE id=?",
            (company_id,),
        ).fetchone()
    company = dict(row) if row else {}

    wb = openpyxl.Workbook()

    # Sheet 1
    ws1 = wb.active
    ws1.title = "종합 평가표"
    _build_summary_sheet(ws1, company, health)

    # Sheet 2
    ws2 = wb.create_sheet("재무비율 분석")
    _build_ratio_sheet(ws2, company_id, company)

    # Sheet 3
    ws3 = wb.create_sheet("재무 원시데이터")
    _build_rawdata_sheet(ws3, company_id, company)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, company.get("company_name", str(company_id))
